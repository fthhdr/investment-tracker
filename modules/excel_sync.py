"""
excel_sync.py — Auto-syncs portfolio data to an Excel workbook.
Called every time the app saves new data (add position, log contribution, etc.)
Creates/updates: investment_tracker/Investment_Tracker.xlsx
"""

import os
import json
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Path config ────────────────────────────────────────────────────────────────
BASE_DIR  = os.path.dirname(os.path.dirname(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "Investment_Tracker.xlsx")

# ── Colour palette (industry standard) ────────────────────────────────────────
CLR = {
    "header_bg":  "1F3864",   # Dark navy — header rows
    "header_fg":  "FFFFFF",   # White text
    "subhdr_bg":  "2E75B6",   # Blue — sub-headers
    "subhdr_fg":  "FFFFFF",
    "alt_row":    "D9E1F2",   # Light blue — alternating rows
    "gain":       "C6EFCE",   # Green background
    "gain_fg":    "276221",
    "loss":       "FFC7CE",   # Red background
    "loss_fg":    "9C0006",
    "input_fg":   "0000FF",   # Blue text = user inputs
    "formula_fg": "000000",   # Black text = formulas
    "total_bg":   "FFF2CC",   # Yellow = totals row
}

def _hdr(ws, row, col, value, bg=None, fg="FFFFFF", bold=True, size=11, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if center:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — PORTFOLIO OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════════
def _write_overview(ws, holdings_df: pd.DataFrame, summary: dict):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 36

    # Title
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "📈  INVESTMENT PORTFOLIO OVERVIEW"
    c.font = Font(name="Arial", bold=True, size=16, color=CLR["header_fg"])
    c.fill = PatternFill("solid", start_color=CLR["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")

    # As-of date
    ws["A2"] = f"Last updated: {datetime.now().strftime('%B %d, %Y  %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=10, color="808080")
    ws.merge_cells("A2:I2")

    # KPI boxes — row 4
    kpis = [
        ("Portfolio Value",  f"${summary.get('total_value', 0):,.2f}"),
        ("Total Invested",   f"${summary.get('total_cost', 0):,.2f}"),
        ("Total P&L",        f"${summary.get('total_pnl', 0):+,.2f}"),
        ("Return %",         f"{summary.get('total_pnl_pct', 0):+.2f}%"),
        ("# Positions",      str(summary.get('num_positions', 0))),
    ]
    for i, (label, val) in enumerate(kpis, start=1):
        col = i * 2 - 1
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
        lbl_cell = ws.cell(row=4, column=col, value=label)
        lbl_cell.font = Font(name="Arial", bold=True, size=9, color="808080")
        lbl_cell.alignment = Alignment(horizontal="center")
        val_cell = ws.cell(row=5, column=col, value=val)
        pnl = summary.get('total_pnl', 0)
        bg = CLR["gain"] if pnl >= 0 else CLR["loss"]
        fg = CLR["gain_fg"] if pnl >= 0 else CLR["loss_fg"]
        if label in ("Total P&L", "Return %"):
            val_cell.fill = PatternFill("solid", start_color=bg)
            val_cell.font = Font(name="Arial", bold=True, size=13, color=fg)
        else:
            val_cell.font = Font(name="Arial", bold=True, size=13)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[5].height = 28

    # Holdings table header — row 7
    headers = ["Symbol", "Name", "Type", "Shares", "Avg Cost ($)",
               "Current Price ($)", "Market Value ($)", "P&L ($)", "P&L (%)", "Allocation (%)"]
    ws.row_dimensions[7].height = 22
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 7, col, h, bg=CLR["subhdr_bg"])

    # Holdings rows
    for i, (_, row) in enumerate(holdings_df.iterrows(), start=8):
        alt = (i % 2 == 0)
        bg_row = CLR["alt_row"] if alt else "FFFFFF"
        pnl_val = row.get("P&L ($)", 0)
        pnl_bg  = CLR["gain"] if pnl_val >= 0 else CLR["loss"]
        pnl_fg  = CLR["gain_fg"] if pnl_val >= 0 else CLR["loss_fg"]

        vals = [
            row["Symbol"], row["Name"], row["Type"], row["Shares"],
            row["Avg Cost"], row["Current Price"], row["Market Value"],
            row["P&L ($)"], row["P&L (%)"] / 100, row.get("Allocation %", 0) / 100,
        ]
        fmts = [
            None, None, None, "0.000000", '$#,##0.00', '$#,##0.00',
            '$#,##0.00', '$#,##0.00;($#,##0.00)', '0.00%', '0.00%',
        ]
        for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = PatternFill("solid", start_color=bg_row)
            cell.font = Font(name="Arial", size=10)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if col > 3 else "left",
                                       vertical="center")
            if fmt:
                cell.number_format = fmt
            # Colour P&L columns
            if col in (8, 9):
                cell.fill = PatternFill("solid", start_color=pnl_bg)
                cell.font = Font(name="Arial", size=10, color=pnl_fg, bold=True)

        ws.row_dimensions[i].height = 18

    # Totals row
    last = 7 + len(holdings_df)
    total_row = last + 1
    ws.row_dimensions[total_row].height = 22
    total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
    total_cell.font = Font(name="Arial", bold=True, size=11)
    total_cell.fill = PatternFill("solid", start_color=CLR["total_bg"])
    # Formula totals for Market Value, P&L
    for col, fmt in [(7, '$#,##0.00'), (8, '$#,##0.00;($#,##0.00)'), (10, '0.00%')]:
        col_l = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({col_l}8:{col_l}{last})")
        cell.number_format = fmt
        cell.font = Font(name="Arial", bold=True, size=11, color=CLR["formula_fg"])
        cell.fill = PatternFill("solid", start_color=CLR["total_bg"])
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, {
        "A": 8, "B": 28, "C": 9, "D": 11, "E": 14,
        "F": 16, "G": 16, "H": 14, "I": 10, "J": 13,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — TRANSACTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def _write_transactions(ws, transactions: list):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "TRANSACTION HISTORY"
    c.font = Font(name="Arial", bold=True, size=14, color=CLR["header_fg"])
    c.fill = PatternFill("solid", start_color=CLR["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Date", "Symbol", "Type", "Asset Class", "Shares", "Price ($)", "Total ($)"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 2, col, h, bg=CLR["subhdr_bg"])

    for i, tx in enumerate(sorted(transactions, key=lambda x: x["date"], reverse=True), start=3):
        alt = (i % 2 == 0)
        bg_row = CLR["alt_row"] if alt else "FFFFFF"
        is_buy = "BUY" in tx.get("type", "")
        type_bg = CLR["gain"] if is_buy else CLR["loss"]
        vals = [
            tx["date"], tx["symbol"], tx["type"],
            tx.get("asset_type", "").upper(),
            tx["shares"], tx["price"], tx["total"],
        ]
        fmts = [None, None, None, None, "0.000000", "$#,##0.00", "$#,##0.00"]
        for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = PatternFill("solid", start_color=type_bg if col == 3 else bg_row)
            cell.font = Font(name="Arial", size=10,
                             color=CLR["gain_fg"] if (col == 3 and is_buy) else
                             CLR["loss_fg"] if col == 3 else "000000")
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if col > 2 else "left",
                                       vertical="center")
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[i].height = 17

    _set_col_widths(ws, {"A": 12, "B": 10, "C": 14, "D": 12, "E": 13, "F": 13, "G": 13})


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — MONTHLY TRACKER
# ═══════════════════════════════════════════════════════════════════════════════
def _write_monthly(ws, contributions: list):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "MONTHLY INVESTMENT TRACKER"
    c.font = Font(name="Arial", bold=True, size=14, color=CLR["header_fg"])
    c.fill = PatternFill("solid", start_color=CLR["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = ["Month", "Planned ($)", "Actual ($)", "Difference ($)",
               "% Complete", "Status", "Notes"]
    for col, h in enumerate(headers, start=1):
        _hdr(ws, 2, col, h, bg=CLR["subhdr_bg"])

    for i, c_data in enumerate(contributions, start=3):
        planned = c_data.get("planned", 0)
        actual  = c_data.get("actual", 0)
        diff    = actual - planned
        pct     = (actual / planned) if planned else 0
        status  = "✅ On Track" if actual >= planned else ("⚠️ Pending" if actual == 0 else "🔴 Under")
        alt = (i % 2 == 0)
        bg_row = CLR["alt_row"] if alt else "FFFFFF"
        status_bg = CLR["gain"] if actual >= planned else (CLR["loss"] if actual > 0 else "FFF2CC")

        vals   = [c_data["month"], planned, actual, diff, pct, status, ""]
        fmts   = [None, "$#,##0.00", "$#,##0.00", "$#,##0.00;($#,##0.00)", "0.0%", None, None]
        for col, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.fill = PatternFill("solid", start_color=status_bg if col == 6 else bg_row)
            cell.font = Font(name="Arial", size=10)
            cell.border = _border()
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                       vertical="center")
            if fmt:
                cell.number_format = fmt
        ws.row_dimensions[i].height = 17

    # Totals
    last = 2 + len(contributions)
    tr = last + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
    for col, fmt in [(2, "$#,##0.00"), (3, "$#,##0.00"), (4, "$#,##0.00;($#,##0.00)")]:
        col_l = get_column_letter(col)
        cell = ws.cell(row=tr, column=col, value=f"=SUM({col_l}3:{col_l}{last})")
        cell.number_format = fmt
        cell.font = Font(name="Arial", bold=True, color=CLR["formula_fg"])
        cell.fill = PatternFill("solid", start_color=CLR["total_bg"])
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center")

    _set_col_widths(ws, {"A": 12, "B": 14, "C": 14, "D": 16, "E": 12, "F": 14, "G": 20})


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN SYNC FUNCTION — call this to rebuild the Excel file
# ═══════════════════════════════════════════════════════════════════════════════
def sync_to_excel(holdings_df=None, summary=None, transactions=None, contributions=None):
    """
    Rebuild Investment_Tracker.xlsx with latest data.
    Safe to call on every app save — overwrites cleanly.
    """
    if holdings_df is None:
        from modules.portfolio import get_enriched_portfolio, get_portfolio_summary
        holdings_df = get_enriched_portfolio()
        summary = get_portfolio_summary(holdings_df)

    if transactions is None:
        from modules.portfolio import load_transactions
        data = load_transactions()
        transactions  = data.get("transactions", [])
        contributions = data.get("monthly_contributions", [])

    wb = Workbook()

    # Sheet 1 — Overview
    ws1 = wb.active
    ws1.title = "📊 Portfolio Overview"
    _write_overview(ws1, holdings_df, summary or {})

    # Sheet 2 — Transactions
    ws2 = wb.create_sheet("📋 Transactions")
    _write_transactions(ws2, transactions or [])

    # Sheet 3 — Monthly Tracker
    ws3 = wb.create_sheet("📅 Monthly Tracker")
    _write_monthly(ws3, contributions or [])

    wb.save(EXCEL_PATH)
    return EXCEL_PATH
